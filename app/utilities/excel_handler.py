from app.utilities import mysql_conn
import xlrd
from openpyxl import load_workbook

from app.utilities.vstd_utils import GetCode


def excel_to_db(excel_path, table: str, columns: list = []):
    excel_sheet =xlrd.open_workbook(excel_path)
    sheet_name = excel_sheet.sheet_names()
    for sh in range(0,len(sheet_name)):
        sheet = excel_sheet.sheet_by_index(sh)
        row_values = ''
        for r in range(1,sheet.nrows):
            row_value = ''
            for i in range(len(columns)):
                row_value = row_value + f'{sheet.cell(r,i).value},'
            row_value = f"({row_value.rstrip(',')})"
            row_values = row_values + ','+row_value
        final_value_set = (row_values.lstrip(','))
    cols =''
    a = final_value_set.replace('(',"('")
    b = a.replace(',',"','")
    c = b.replace(")','(",'),(')
    d = c.replace(")","')")
    e = d.replace('.0','')
    for i in range(len(columns)):
        if i > 0 :
            act_col = ',' + columns[i]
        else :
            act_col = columns[i]
        cols = cols + act_col
    query = f"INSERT INTO {table} ({cols}) VALUES {e}"
    # print(query)
    m_conn = mysql_conn.mysql_obj()
    m_conn.mysql_execute(query, fetch_result=False)
    m_conn.commit()
    m_conn.close()
    return True


def add_questions_from_xl(excel_path):
    wb = load_workbook(filename=excel_path)
    getCode = GetCode()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for key, *values in ws.iter_rows(min_row=2):
            if key.value == None:
                break
            q_content = key.value
            str_insert = ''
            i = 0
            for v in values:
                item = v.value
                if item == None:
                    break
                if i == 6:
                    item = getCode.subject(v.value)
                i += 1
                str_insert = str_insert + f"'{str(item)}'" + ','
            query = f"INSERT INTO mcqs (q_id,content,opt1,opt2,opt3,opt4,ans,class,sub_id,added_by) VALUES (NULL,'{q_content}',{str_insert[:-1]},'{sheet}')"
            m_conn = mysql_conn.mysql_obj()
            m_conn.mysql_execute(query, fetch_result=False)
            m_conn.commit()
            m_conn.close()
            break
    return True

